"""
Monterosa Succulents - weekly availability catalog generator.

Renders a print-ready PDF catalog (3 columns x 4 rows per page) from a list
of (image, variety name, pot size) items. Uses direct PIL pixel rendering
(NOT html->pdf) so layout is exact and never crops photos.
"""

import io
import re
import datetime
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Optional, List

from PIL import Image, ImageDraw, ImageFont, ImageOps

import os as _os
FDIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "fonts") + "/"

# ---------- geometry (mm) ----------
DPI = 200
MM = DPI / 25.4

PAGE_W = round(210 * MM)
PAGE_H = round(297 * MM)

MARGIN_L = round(12 * MM)
MARGIN_R = round(12 * MM)
MARGIN_T = round(12 * MM)
MARGIN_B = round(10 * MM)
HEADER_H = round(24 * MM)
FOOTER_H = round(10 * MM)

COLS, ROWS = 3, 4
PER_PAGE = COLS * ROWS
COL_GAP = round(4.5 * MM)
ROW_GAP = round(5 * MM)
LABEL_H = round(11.5 * MM)
CARD_PAD = round(2.5 * MM)
TEXT_TOP_GAP = round(2.8 * MM)

content_w = PAGE_W - MARGIN_L - MARGIN_R
content_h = PAGE_H - MARGIN_T - MARGIN_B - HEADER_H - FOOTER_H
col_w = (content_w - COL_GAP * (COLS - 1)) // COLS
row_h = (content_h - ROW_GAP * (ROWS - 1)) // ROWS
photo_w = col_w - 2 * CARD_PAD
photo_h = row_h - LABEL_H - 2 * CARD_PAD

# ---------- colors ----------
OLIVE = (134, 127, 69)
DARK_TEXT = (58, 54, 36)
MUTED = (138, 133, 112)
CARD_BG = (253, 252, 249)
PHOTO_BG = (242, 240, 230)
CARD_BORDER = (230, 226, 211)
BADGE_BG = (240, 238, 224)
WHITE = (255, 255, 255)

# ---------- fonts ----------
f_title = ImageFont.truetype(FDIR + "IBMPlexSerif-Bold.ttf", round(6.4 * MM))
f_name = ImageFont.truetype(FDIR + "CrimsonPro-Italic.ttf", round(4.0 * MM))
f_pot = ImageFont.truetype(FDIR + "InstrumentSans-Bold.ttf", round(2.9 * MM))
f_footer = ImageFont.truetype(FDIR + "InstrumentSans-Regular.ttf", round(3.0 * MM))

TEXT = {
    "en": {
        "header": "Monterosa Succulents Availability",
        "footer": "Page {page} of {total}  \u00b7  Week {week}, {year}",
    },
    "pt": {
        "header": "Disponibilidade de Suculentas \u00b7 Monterosa",
        "footer": "P\u00e1gina {page} de {total}  \u00b7  Semana {week}, {year}",
    },
}


@dataclass
class CatalogItem:
    image: "Image.Image"   # PIL Image, already loaded
    name: str
    pot_cm: Optional[str] = None


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_name_and_pot(raw_name: str):
    """'Crassula rogersii P14' -> ('Crassula rogersii', '14')"""
    m = re.match(r"(.+?)\s+P(\d+)$", raw_name.strip())
    if m:
        return m.group(1), m.group(2)
    return raw_name.strip(), None


def name_from_filename(filename: str):
    """
    Derive a display name + pot size from an uploaded photo's filename.
    Examples:
      'Aloe_vera_P14.jpg'          -> ('Aloe vera', '14')
      'Echeveria-Blue-Prince.png'  -> ('Echeveria Blue Prince', None)
      'Kalanchoe tomentosa Nigra 14cm.jpg' -> ('Kalanchoe tomentosa Nigra', '14')
    """
    stem = re.sub(r"\.[A-Za-z0-9]+$", "", filename)
    stem = stem.replace("_", " ").replace("-", " ")
    stem = re.sub(r"\s+", " ", stem).strip()

    pot = None
    m = re.search(r"\bP(\d{1,2})\b$", stem, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"\b(\d{1,2})\s*cm\b$", stem, flags=re.IGNORECASE)
    if m:
        pot = m.group(1)
        stem = stem[:m.start()].strip()

    return stem, pot


def extract_items_from_xlsx(xlsx_bytes: bytes) -> List[CatalogItem]:
    """
    Extract (name, pot, image) triples from a Monterosa-style xlsx: column B
    has 'Variety Name P14' text, column C has one embedded photo per row.
    Photo files are not always numbered in row order in the raw xlsx, so we
    reconstruct the true top-to-bottom order from the drawing anchors
    (row + vertical offset) rather than trusting the image filenames.
    """
    import openpyxl

    buf = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    names = []
    r = 3
    while ws.cell(row=r, column=2).value:
        names.append(ws.cell(row=r, column=2).value)
        r += 1

    ns = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }
    z = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    drawing_paths = [n for n in z.namelist() if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
    drawing_path = drawing_paths[0]
    rels_path = f"xl/drawings/_rels/{drawing_path.split('/')[-1]}.rels"

    rels_root = ET.fromstring(z.read(rels_path))
    rid_to_target = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels_root}

    drawing_root = ET.fromstring(z.read(drawing_path))
    anchors = []
    for anchor in drawing_root:
        frm = anchor.find('xdr:from', ns)
        row = int(frm.find('xdr:row', ns).text)
        rowoff = int(frm.find('xdr:rowOff', ns).text)
        blip = anchor.find('.//a:blip', ns)
        rid = blip.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
        target = rid_to_target[rid].split('/')[-1]
        anchors.append((row, rowoff, target))

    anchors.sort(key=lambda a: (a[0], a[1]))
    image_order = [a[2] for a in anchors]

    items = []
    for name, img_file in zip(names, image_order):
        clean_name, pot = parse_name_and_pot(name)
        img_bytes = z.read(f"xl/media/{img_file}")
        im = Image.open(io.BytesIO(img_bytes))
        im = ImageOps.exif_transpose(im).convert("RGB")
        items.append(CatalogItem(image=im, name=clean_name, pot_cm=pot))
    return items


def build_items_from_photos(uploaded_files) -> List[CatalogItem]:
    """uploaded_files: iterable of (filename, bytes)"""
    items = []
    for filename, data in uploaded_files:
        name, pot = name_from_filename(filename)
        im = Image.open(io.BytesIO(data))
        im = ImageOps.exif_transpose(im).convert("RGB")
        items.append(CatalogItem(image=im, name=name, pot_cm=pot))
    return items


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def draw_rounded_rect(draw, box, radius, fill=None, outline=None, width=1):
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def text_w(draw, text, font):
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def paste_contain(canvas, im, box):
    x0, y0, x1, y1 = box
    bw, bh = x1 - x0, y1 - y0
    iw, ih = im.size
    scale = min(bw / iw, bh / ih)
    nw, nh = max(1, round(iw * scale)), max(1, round(ih * scale))
    resized = im.resize((nw, nh), Image.LANCZOS)
    px = x0 + (bw - nw) // 2
    py = y0 + (bh - nh) // 2
    canvas.paste(resized, (px, py))


def render_header(canvas, draw, logo, header_text):
    y0 = MARGIN_T
    line_y = y0 + HEADER_H - round(4 * MM)

    logo_h = round(15.5 * MM)
    scale = logo_h / logo.height
    logo_resized = logo.resize((round(logo.width * scale), logo_h), Image.LANCZOS)
    logo_y = y0 + (HEADER_H - round(4 * MM) - logo_h) // 2
    canvas.paste(logo_resized, (MARGIN_L, logo_y), logo_resized if logo_resized.mode == "RGBA" else None)

    tw = text_w(draw, header_text, f_title)
    th_bbox = draw.textbbox((0, 0), header_text, font=f_title)
    th = th_bbox[3] - th_bbox[1]
    right_x = PAGE_W - MARGIN_R
    title_y = y0 + (HEADER_H - round(4 * MM) - th) // 2 - th_bbox[1]
    draw.text((right_x - tw, title_y), header_text, font=f_title, fill=OLIVE)

    draw.line([(MARGIN_L, line_y), (PAGE_W - MARGIN_R, line_y)], fill=OLIVE, width=round(0.6 * MM))


def render_footer(canvas, draw, footer_text):
    tw = text_w(draw, footer_text, f_footer)
    y = PAGE_H - MARGIN_B - round(3.2 * MM)
    draw.text(((PAGE_W - tw) / 2, y), footer_text, font=f_footer, fill=MUTED)


def render_card(canvas, draw, it: CatalogItem, x0, y0):
    x1, y1 = x0 + col_w, y0 + row_h
    draw_rounded_rect(draw, (x0, y0, x1, y1), radius=round(1.2 * MM), fill=CARD_BG, outline=CARD_BORDER, width=1)

    px0, py0 = x0 + CARD_PAD, y0 + CARD_PAD
    px1, py1 = x1 - CARD_PAD, y0 + CARD_PAD + photo_h
    draw_rounded_rect(draw, (px0, py0, px1, py1), radius=round(1.0 * MM), fill=PHOTO_BG)
    paste_contain(canvas, it.image, (px0 + 1, py0 + 1, px1 - 1, py1 - 1))

    label_top = py1 + TEXT_TOP_GAP
    cx = (x0 + x1) / 2
    name_w = text_w(draw, it.name, f_name)
    draw.text((cx - name_w / 2, label_top), it.name, font=f_name, fill=DARK_TEXT)
    ly = label_top + round(3.9 * MM)

    if it.pot_cm:
        badge_txt = f"\u00d8 {it.pot_cm} cm"
        bw = text_w(draw, badge_txt, f_pot)
        pad_x, pad_y = round(2.0 * MM), round(0.8 * MM)
        bx0 = cx - bw / 2 - pad_x
        bx1 = cx + bw / 2 + pad_x
        by0 = ly + round(0.5 * MM)
        by1 = by0 + round(2.9 * MM) + 2 * pad_y
        draw_rounded_rect(draw, (bx0, by0, bx1, by1), radius=round(1.5 * MM), fill=BADGE_BG)
        draw.text((cx - bw / 2, by0 + pad_y), badge_txt, font=f_pot, fill=OLIVE)


def render_page(page_items, logo, header_text, footer_text):
    canvas = Image.new("RGB", (PAGE_W, PAGE_H), WHITE)
    draw = ImageDraw.Draw(canvas)
    render_header(canvas, draw, logo, header_text)
    render_footer(canvas, draw, footer_text)

    grid_y0 = MARGIN_T + HEADER_H
    for idx, it in enumerate(page_items):
        r, c = divmod(idx, COLS)
        x0 = MARGIN_L + c * (col_w + COL_GAP)
        y0 = grid_y0 + r * (row_h + ROW_GAP)
        render_card(canvas, draw, it, x0, y0)
    return canvas


def generate_catalog_pdf(items: List[CatalogItem], logo: "Image.Image", lang: str = "en", sort_alphabetically: bool = True) -> bytes:
    """Returns PDF bytes for the full catalog."""
    lang = lang if lang in TEXT else "en"
    strings = TEXT[lang]
    iso = datetime.date.today().isocalendar()
    week, year = iso[1], iso[0]

    if sort_alphabetically:
        items = sorted(items, key=lambda it: it.name.strip().lower())

    pages_items = [items[i:i + PER_PAGE] for i in range(0, len(items), PER_PAGE)]
    total = len(pages_items)
    page_images = []
    for pi, page_items in enumerate(pages_items):
        footer_text = strings["footer"].format(page=pi + 1, total=total, week=week, year=year)
        img = render_page(page_items, logo, strings["header"], footer_text)
        page_images.append(img)

    out = io.BytesIO()
    page_images[0].save(out, format="PDF", save_all=True, append_images=page_images[1:], resolution=DPI)
    return out.getvalue()
